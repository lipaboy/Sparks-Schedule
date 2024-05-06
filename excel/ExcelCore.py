import openpyxl
import datetime
import copy
import os
from search.SparksScheduleSearch import SparksScheduleSearch
from search.WeekScheduleExcelType import WeekScheduleExcelType, EmployeeCard, ShiftType, TruckDistributionType, TruckElem

FILENAME_SCHEDULE_DATA_BASE = "ScheduleDataBase.xlsx"
FILENAME_POOL_TIMETABLE = "PoolTimetable.xlsx"
WEEK_LENGTH = 7
SPACE_BETWEEN_TABLES = 1
NUMBER_OF_TABLES_IN_LINE = 3
NUMBER_OF_EMPTY_BORDER_CELLS = 10
STARTING_POINT_SCHEDULE_DATA_BASE = {#Coordinates of workspace in SCHEDULE_DATA_BASE
    "List DB": {
        "row": 3,
        "column": 1
    },
    "List STAFF": {
        "STAFF_DATA": {
            "width": 5,
            "row": 3,
            "column": 1
        },
        "PREFER_OF_WEEK": {
            "width": WEEK_LENGTH,
            "row": 3,
            "column": 7
        },
        "CURRENT_WEEK": {
            "width": WEEK_LENGTH+1,
            "row": 2,
            "column": 16
        }
    },
    "List COEFFICIENTS": {
        "row": 1,
        "column": 1
    }
}
STARTING_POINT_POOL_TIMETABLE = {#Coordinates of workspace in POOL_TIMETABLE
    "row": 3,
    "column": 1
}

TABLE_0_COLUMN_WIDTH = 10 #For BD sheet min value 10
TABLE_1_COLUMN_WIDTH = 15 #For Staff and Trucks sheets

FONT_SIZE = 14
THIN_BORDER = openpyxl.styles.Side(border_style="thin", color="000000")
MEDIUM_BORDER = openpyxl.styles.Side(border_style="medium", color="000000")
THICK_BORDER = openpyxl.styles.Side(border_style="thick", color="000000")
COLOR_HALL = "F9E79F"
COLOR_TRUCK = "ABEBC6"
CHAR_CROSS = '✕' #'✖'
CHAR_HALL = 'С' #RUS С
CHAR_HALF_HALL = 'С2' #RUS С
CHAR_TRUCK = 'Т' #RUS Т
ERROR_STR_HEAD = "\n\t**ERROR in excel.ExcelCore"#! (numOfFunction)\n\t\t

def formatting_cell(sheet, row, column, value, fontSize, fontName, fontBold, fontItalic, alignmentHorizontal, alignmentVertical):
    sheet.cell(row=row, column=column).value = value
    sheet.cell(row=row, column=column).font = openpyxl.styles.Font(size=fontSize, name=fontName, bold=fontBold, italic=fontItalic)
    sheet.cell(row=row, column=column).alignment = openpyxl.styles.Alignment(horizontal=alignmentHorizontal, vertical=alignmentVertical)

def get_dated_week(mode=0, currentDay: datetime.date = datetime.date.today()): #it gives dated week, that begin next monday
    if mode == 0:
        months = {
        1: 'Янв.',
        2: 'Февр.',
        3: 'Март',
        4: 'Апр.',
        5: 'Май',
        6: 'Июнь',
        7: 'Июль',
        8: 'Авг.',
        9: 'Сент.',
        10: 'Окт.',
        11: 'Нояб.',
        12: 'Дек.'}
        delta_day = datetime.timedelta(days=1)
        current_date = currentDay #current_date = datetime.date(2024, 11, 14)
        current_date -= current_date.weekday() * delta_day #It gives current week that begin on Monday
        # current_date += (WEEK_LENGTH-current_date.weekday()) * delta_day #It gives next week that begin on Monday
        datedWeek = list()
        for i in range(WEEK_LENGTH):
            datedWeek.append(f"{current_date.day} {months[current_date.month]}")
            current_date += delta_day
        return datedWeek
    elif mode == 1:
        daysWeek = {
            1: 'ПН',
            2: 'ВТ',
            3: "СР",
            4: "ЧТ",
            5: "ПТ",
            6: "СБ",
            7: "ВС"}
        return list(daysWeek.values())
    else:
        print(ERROR_STR_HEAD + "! (get_dated_week)\n\t\tThe wrong mode!")
        return None

def init_schedule_data_base(filenameSceduleDataBase):
    # Checking for the existence of DB's file
    # try:
    #     openpyxl.load_workbook(filename = filenameSceduleDataBase)
    # except:
    #     pass
    # else:
    #     return 0

                    #Init List DB
    wbSceduleDataBase = openpyxl.Workbook()
    sheet = wbSceduleDataBase.worksheets[0]
    sheet.title = "БД расписаний"
    columnCursor = 1
    formatting_cell(sheet, 1, columnCursor, 0, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    columnCursor += 2
    formatting_cell(sheet, 1, columnCursor, "Легенда", FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, 1, columnCursor, CHAR_HALL, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=1, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_HALL)
    sheet.cell(row=1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, 1, columnCursor, CHAR_HALF_HALL, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=1, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_HALL)
    sheet.cell(row=1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, 1, columnCursor, CHAR_TRUCK, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=1, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_TRUCK)
    sheet.cell(row=1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, 1, columnCursor, CHAR_CROSS, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)

                    #Init List STAFF
    startingPointColumn = 1
    headStartingPointRow = 1
    wbSceduleDataBase.create_sheet("Персонал")
    sheet = wbSceduleDataBase.worksheets[1]

                #Subtable STAFF_DATA
    rowCursor = headStartingPointRow
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["column"]
    lengthDataStaff = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["width"]
            #Head
    sheet.merge_cells(start_row=rowCursor, start_column=columnCursor, end_row=rowCursor, end_column=columnCursor + lengthDataStaff-1)
    formatting_cell(sheet, rowCursor, columnCursor, "Данные персонала", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(left=THICK_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor+lengthDataStaff-1).border = openpyxl.styles.Border(right=THICK_BORDER)
        #Subhead
    rowCursor += 1
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["column"]
    # Cell NAME
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "Имя", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(left=MEDIUM_BORDER, right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    # Cell ELDER
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "Старший", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    # Cell NUM_OF_TRUCK (T)
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "Кол-во Т", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    # Cell NUM_OF_HALL (C)
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "Кол-во С", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    # Cell PREFER_NUM_OF_HALL (C)
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH * 2
    formatting_cell(sheet, rowCursor, columnCursor, "Желаемое кол-во С", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=THICK_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)

                #Subtable PREFER_OF_WEEK
    rowCursor = headStartingPointRow
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["column"]
            #Head Subtable PREFER_OF_WEEK
    sheet.merge_cells(start_row=rowCursor, start_column=columnCursor, end_row=rowCursor, end_column=columnCursor + WEEK_LENGTH-1)
    formatting_cell(sheet, rowCursor, columnCursor, "Предпочтения на неделю", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(left=THICK_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor+WEEK_LENGTH-1).border = openpyxl.styles.Border(right=THICK_BORDER)
        #Legend
    columnCursor += WEEK_LENGTH
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "Легенда", FONT_SIZE, "Times New Roman", True, True, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=MEDIUM_BORDER, top=MEDIUM_BORDER)
    formatting_cell(sheet, rowCursor + 1, columnCursor, CHAR_CROSS, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=rowCursor+1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
        #Subhead (days of week)
    rowCursor += 1
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["column"]
    datedWeek = get_dated_week(1)
    for i in range(WEEK_LENGTH):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor + i)].width = TABLE_0_COLUMN_WIDTH
        formatting_cell(sheet, rowCursor, columnCursor + i, datedWeek[i], FONT_SIZE, "Times New Roman", True, False, "center", "center")
        sheet.cell(row=rowCursor, column=columnCursor + i).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(left=THICK_BORDER, right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor+WEEK_LENGTH-1).border = openpyxl.styles.Border(left=MEDIUM_BORDER, right=THICK_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)

                #Subtable CURRENT_WEEK
    rowCursor = headStartingPointRow
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["column"]
            #Head Subtable CURRENT_WEEK
    sheet.merge_cells(start_row=rowCursor, start_column=columnCursor, end_row=rowCursor, end_column=columnCursor + WEEK_LENGTH)
    formatting_cell(sheet, rowCursor, columnCursor, "Текущая неделя", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(left=THICK_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor + WEEK_LENGTH).border = openpyxl.styles.Border(right=THICK_BORDER)
        #Subhead days of week
    rowCursor += 1
    columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["column"]
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, rowCursor, columnCursor, "", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=rowCursor, column=columnCursor).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    columnCursor += 1
    for i in range(WEEK_LENGTH):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor + i)].width = TABLE_0_COLUMN_WIDTH
        formatting_cell(sheet, rowCursor, columnCursor + i, datedWeek[i], FONT_SIZE, "Times New Roman", True, False, "center", "center")
        sheet.cell(row=rowCursor, column=columnCursor + i).border = openpyxl.styles.Border(right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor-1).border = openpyxl.styles.Border(left=THICK_BORDER, right=MEDIUM_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)
    sheet.cell(row=rowCursor, column=columnCursor+WEEK_LENGTH-1).border = openpyxl.styles.Border(left=MEDIUM_BORDER, right=THICK_BORDER, bottom=THICK_BORDER, top=MEDIUM_BORDER)

    # Creating border of empty cells
    rowCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["row"]
    for j in range(NUMBER_OF_EMPTY_BORDER_CELLS):
        columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["column"]
        for i in range(STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["width"]):
            sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, bottom=THIN_BORDER)
            columnCursor += 1
        sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER)

        columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["column"]
        for i in range(STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["width"]):
            sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, bottom=THIN_BORDER)
            columnCursor += 1
        sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER)

        columnCursor = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["column"]
        for i in range(STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["width"]):
            sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, bottom=THIN_BORDER)
            columnCursor += 1
        sheet.cell(row=rowCursor + j, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER)

                    #Init List COEFFICIENTS
    startingPointColumn = 1
    headStartingPointRow = 1
    wbSceduleDataBase.create_sheet("Коэффиценты")
    sheet = wbSceduleDataBase.worksheets[2]
    # """ Коэффициент дебатов для непрерывный череды смен""" self.shiftRepeatCoef = 10
    # """ Коэффициент дебатов для разницы между реальным количеством смен и желаемым для сотрудника""" self.differInShiftsCoef = 2.1
    # """ Коэффициент дебатов, когда сотрудник выходит на смену в не желаемый день""" self.undesirableDayCoef = 4
    #Cell shiftRepeatCoef
    columnCursor = startingPointColumn
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = 2 * TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, headStartingPointRow, columnCursor, "shiftRepeatCoef", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=headStartingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THICK_BORDER, right=THICK_BORDER, bottom=THICK_BORDER)
    formatting_cell(sheet, headStartingPointRow + 1, columnCursor, 10, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=headStartingPointRow+1, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER)
    #Cell differInShiftsCoef
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = 2 * TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, headStartingPointRow, columnCursor, "differInShiftsCoef", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=headStartingPointRow, column=columnCursor).border = openpyxl.styles.Border(right=THICK_BORDER, bottom=THICK_BORDER)
    formatting_cell(sheet, headStartingPointRow + 1, columnCursor, 2.1, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=headStartingPointRow+1, column=columnCursor).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)
    #Cell undesirableDayCoef
    columnCursor += 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = 2 * TABLE_1_COLUMN_WIDTH
    formatting_cell(sheet, headStartingPointRow, columnCursor, "undesirableDayCoef", FONT_SIZE, "Times New Roman", True, False, "center", "center")
    sheet.cell(row=headStartingPointRow, column=columnCursor).border = openpyxl.styles.Border(right=THICK_BORDER, bottom=THICK_BORDER)
    formatting_cell(sheet, headStartingPointRow + 1, columnCursor, 4, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=headStartingPointRow+1, column=columnCursor).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)

    #Save
    wbSceduleDataBase.save(filenameSceduleDataBase)
    print(f"\n\tFile '{filenameSceduleDataBase}' was created!")
    return 1

def output_pool_of_schedule_to_excel(filenameSceduleDataBase, filenamePoolTimetable, searchMode="fast", currentDay: datetime.date = datetime.date.today()):# "fast", "part", "full"
    #Init DB's FILE
    init_schedule_data_base(filenameSceduleDataBase)
    #Init staff(elders, ghosts)
    staff = get_schedule_list_staff(filenameSceduleDataBase, 1)
    elders = list()
    ghosts = list()
    for staffName, staffPosition in staff.items():
        if staffPosition == True:
            elders.append(staffName)
        else:
            ghosts.append(staffName)
    #Init
    sparks = SparksScheduleSearch()
    _prevSchedule = get_schedule_list_data_base(filenameSceduleDataBase)
    if _prevSchedule != None:
        _prevSchedule.Trucks = get_truck_distribution_data_base(filenameSceduleDataBase)
    coefs = get_schedule_list_coefficients(filenameSceduleDataBase)
    timeTable = sparks.search(eldermen=elders,
                              ghostmen=ghosts,
                              undesirableDays=get_schedule_list_staff(filenameSceduleDataBase, 5),
                              shiftCountPreferences=get_schedule_list_staff(filenameSceduleDataBase, 4),
                              prevSchedule=_prevSchedule,
                              coeficiencies=coefs,
                              mode=searchMode) #'fast', 'part', 'full'
    wb = openpyxl.Workbook()
    sheet = wb.worksheets[0]
    sheet.title = "Выбор расписания"
    tableWidth = 1 + WEEK_LENGTH + SPACE_BETWEEN_TABLES  # +1 - begin from 1; +1 - space between tables
    tableHeight = 1 + len(timeTable[0].EmployeeCards) + SPACE_BETWEEN_TABLES
    #Legend
    startingPointRow = 1
    columnCursor = 1
    sheet.column_dimensions[openpyxl.utils.get_column_letter(columnCursor)].width = TABLE_0_COLUMN_WIDTH
    formatting_cell(sheet, startingPointRow, columnCursor, "Легенда", FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=startingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, startingPointRow, columnCursor, CHAR_HALL, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=startingPointRow, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_HALL)
    sheet.cell(row=startingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, startingPointRow, columnCursor, CHAR_HALF_HALL, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=startingPointRow, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_HALL)
    sheet.cell(row=startingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, startingPointRow, columnCursor, CHAR_TRUCK, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=startingPointRow, column=columnCursor).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor=COLOR_TRUCK)
    sheet.cell(row=startingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    columnCursor += 1
    formatting_cell(sheet, startingPointRow, columnCursor, CHAR_CROSS, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    sheet.cell(row=startingPointRow, column=columnCursor).border = openpyxl.styles.Border(left=THIN_BORDER, right=THIN_BORDER, bottom=THIN_BORDER, top=THIN_BORDER)
    #Fill table all timeTables
    for numOfTable in range(len(timeTable)):
        startingPointColumn = 1 + ((numOfTable % NUMBER_OF_TABLES_IN_LINE) * tableWidth)
        startingPointRow = 3 + ((numOfTable // NUMBER_OF_TABLES_IN_LINE) * tableHeight)
        sheet.column_dimensions[openpyxl.utils.get_column_letter(startingPointColumn)].width = TABLE_0_COLUMN_WIDTH
        for i in range(1, WEEK_LENGTH + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(startingPointColumn+i)].width = TABLE_0_COLUMN_WIDTH#15 #if [day, dd mmmm]
        formatting_cell(sheet, startingPointRow, startingPointColumn, f"№ {numOfTable + 1}", FONT_SIZE, "Times New Roman", False, False, "center", "center")
        #Formatting and x-filling the timetable
        for row in range(startingPointRow+1, len(timeTable[numOfTable].EmployeeCards)+startingPointRow+1):
            for column in range(startingPointColumn+1, WEEK_LENGTH + startingPointColumn + 1):
                sheet.cell(row=row, column=column).border = openpyxl.styles.Border(right=THIN_BORDER,bottom=THIN_BORDER)
                formatting_cell(sheet, row, column, CHAR_CROSS, FONT_SIZE, "Times New Roman", False, False, "center", "center")
        #Creating of header of timetable
        datedWeek = get_dated_week(currentDay=currentDay)
        for day in range(1, WEEK_LENGTH + 1):
            sheet.cell(row=startingPointRow, column=startingPointColumn+day).border = openpyxl.styles.Border(left=THICK_BORDER, right=THICK_BORDER, top=THICK_BORDER, bottom=THICK_BORDER)
            formatting_cell(sheet, startingPointRow, startingPointColumn+day, datedWeek[day-1], 12, "Times New Roman", True, False, "center", "center")
        for numOfStaff in range(len(timeTable[numOfTable].EmployeeCards)):
            sheet.cell(row=startingPointRow + numOfStaff + 1, column=startingPointColumn).border = openpyxl.styles.Border(left=THICK_BORDER, right=THICK_BORDER, top=THICK_BORDER, bottom=THICK_BORDER)
            formatting_cell(sheet, startingPointRow + numOfStaff + 1, startingPointColumn, timeTable[numOfTable].EmployeeCards[numOfStaff].Name, FONT_SIZE, "Times New Roman", timeTable[numOfTable].EmployeeCards[numOfStaff].IsElder, True, "right", "center")
        #Fill data to the timetable
        typeOfShift = CHAR_CROSS
        for numOfStaff in range(len(timeTable[numOfTable].EmployeeCards)):  #class EmployeeCard(name: str, isElder: bool, shifts: list[ShiftType])
            for shift in timeTable[numOfTable].EmployeeCards[numOfStaff].Shifts: #ShiftType = tuple[DayType, ShiftLength, PlaceToWork]
                if shift[2] == 'Hall':
                    sheet.cell(row=startingPointRow + numOfStaff + 1, column=startingPointColumn + shift[0]).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='F9E79F')
                    if shift[1] == 1:
                        typeOfShift = CHAR_HALL
                    else:
                        typeOfShift = CHAR_HALF_HALL
                else:
                    sheet.cell(row=startingPointRow + numOfStaff + 1, column=startingPointColumn + shift[0]).fill = openpyxl.styles.PatternFill(fill_type='solid', fgColor='ABEBC6')
                    typeOfShift = CHAR_TRUCK
                formatting_cell(sheet, startingPointRow + numOfStaff + 1, startingPointColumn + shift[0], typeOfShift, FONT_SIZE, "Times New Roman", False, False, "center", "center")
    #Close and save excel's file
    wb.save(filenamePoolTimetable)
    print(f"\n\tFile '{filenamePoolTimetable}' was created!")
    #Return number of tables in pool
    lengthOfPool = len(timeTable)
    return lengthOfPool

def update_schedule_data_base(filenameSceduleDataBase, filenamePoolTimetable, numChoosingTimetable):
    #Init
    wbSceduleDataBase = openpyxl.load_workbook(filename = filenameSceduleDataBase)
    sheetDB = wbSceduleDataBase.worksheets[0]
    sheetSTAFF = wbSceduleDataBase.worksheets[1]
    wbChooseTimetable = openpyxl.load_workbook(filename = filenamePoolTimetable)
    poolSheet = wbChooseTimetable.worksheets[0]
    #Calc of staff's size
    staffLength = 0
    r = 4
    while poolSheet.cell(row=r, column=1).value != None:
        staffLength += 1
        r += 1
    #Init tables parameters
    tableWidth = 1 + WEEK_LENGTH + SPACE_BETWEEN_TABLES  # +1 - begin from 1; 7 - week length; +1 - space between tables
    tableHeight = 1 + staffLength + SPACE_BETWEEN_TABLES
    startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["column"] + (sheetDB.cell(row=1, column=1).value * tableWidth)
    startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["row"]
    sheetSTAFF_startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["column"]
    sheetSTAFF_startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["CURRENT_WEEK"]["row"]
    poolPointColumn = STARTING_POINT_POOL_TIMETABLE["column"] + (((numChoosingTimetable-1) % NUMBER_OF_TABLES_IN_LINE) * tableWidth)
    poolPointRow = STARTING_POINT_POOL_TIMETABLE["row"] + (((numChoosingTimetable-1) // NUMBER_OF_TABLES_IN_LINE) * tableHeight)
    #Update List DB and List STAFF CURRENT_WEEK
    for j in range(tableWidth-SPACE_BETWEEN_TABLES):#Head update DB
        sheetDB.column_dimensions[openpyxl.utils.get_column_letter(startingPointColumn+j)].width = TABLE_0_COLUMN_WIDTH
        sheetDB.cell(row=startingPointRow, column=startingPointColumn + j).value = poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).value
        sheetDB.cell(row=startingPointRow, column=startingPointColumn + j).font = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).font)
        sheetDB.cell(row=startingPointRow, column=startingPointColumn + j).alignment = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).alignment)
        sheetDB.cell(row=startingPointRow, column=startingPointColumn + j).border = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).border)
        sheetDB.cell(row=startingPointRow, column=startingPointColumn + j).fill = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).fill)
    for j in range(tableWidth - SPACE_BETWEEN_TABLES):  # Head update STAFF CURRENT_WEEK
        sheetSTAFF.column_dimensions[openpyxl.utils.get_column_letter(startingPointColumn + j)].width = TABLE_0_COLUMN_WIDTH
        sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn + j).value = poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).value
        sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn + j).font = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).font)
        sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn + j).alignment = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).alignment)
        sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn + j).border = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).border)
        sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn + j).fill = copy.copy(poolSheet.cell(row=poolPointRow, column=poolPointColumn + j).fill)
    poolTable = WeekScheduleExcelType() #For calcNewTrucks()
    for i in range(1, tableHeight-SPACE_BETWEEN_TABLES):#Other
        bufferName = poolSheet.cell(row=poolPointRow + i, column=poolPointColumn).value#For calcNewTrucks()
        bufferIsElder = poolSheet.cell(row=poolPointRow + i, column=poolPointColumn).font.b#For calcNewTrucks()
        bufferShifts = list[ShiftType]()#For calcNewTrucks()
        for j in range(tableWidth-SPACE_BETWEEN_TABLES):
            if poolSheet.cell(row=poolPointRow+i, column=poolPointColumn+j).value != CHAR_CROSS:#For calcNewTrucks()
                if poolSheet.cell(row=poolPointRow+i, column=poolPointColumn+j).value.upper() == CHAR_TRUCK:#For calcNewTrucks()
                    bufferShifts.append((j, 1.0, "Truck"))#For calcNewTrucks()
                elif poolSheet.cell(row=poolPointRow+i, column=poolPointColumn+j).value.upper() == CHAR_HALL:#For calcNewTrucks()
                    bufferShifts.append((j, 1.0, "Hall"))#For calcNewTrucks()
                elif poolSheet.cell(row=poolPointRow+i, column=poolPointColumn+j).value.upper() == CHAR_HALF_HALL:#For calcNewTrucks()
                    bufferShifts.append((j, 0.5, "Hall"))#For calcNewTrucks()
            sheetDB.cell(row=startingPointRow + i, column=startingPointColumn + j).value = poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).value
            sheetDB.cell(row=startingPointRow + i, column=startingPointColumn + j).font = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).font)
            sheetDB.cell(row=startingPointRow + i, column=startingPointColumn + j).alignment = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).alignment)
            sheetDB.cell(row=startingPointRow + i, column=startingPointColumn + j).border = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).border)
            sheetDB.cell(row=startingPointRow + i, column=startingPointColumn + j).fill = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).fill)
            sheetSTAFF.cell(row=sheetSTAFF_startingPointRow + i, column=sheetSTAFF_startingPointColumn + j).value = poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).value
            sheetSTAFF.cell(row=sheetSTAFF_startingPointRow + i, column=sheetSTAFF_startingPointColumn + j).font = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).font)
            sheetSTAFF.cell(row=sheetSTAFF_startingPointRow + i, column=sheetSTAFF_startingPointColumn + j).alignment = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).alignment)
            sheetSTAFF.cell(row=sheetSTAFF_startingPointRow + i, column=sheetSTAFF_startingPointColumn + j).border = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).border)
            sheetSTAFF.cell(row=sheetSTAFF_startingPointRow + i, column=sheetSTAFF_startingPointColumn + j).fill = copy.copy(poolSheet.cell(row=poolPointRow + i, column=poolPointColumn + j).fill)
        poolTable.EmployeeCards.append(EmployeeCard(name=bufferName, isElder=bufferIsElder, shifts=bufferShifts))#For calcNewTrucks()
    sheetDB.cell(row=1, column=1).value = sheetDB.cell(row=1, column=1).value + 1  # number of week schedule + 1 after added
    sheetDB.cell(row=startingPointRow, column=startingPointColumn).value = "№ " + str(sheetDB.cell(row=1, column=1).value)
    sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn).value = sheetDB.cell(row=startingPointRow, column=startingPointColumn).value
    sheetSTAFF.cell(row=sheetSTAFF_startingPointRow, column=sheetSTAFF_startingPointColumn).border = openpyxl.styles.Border(top= THICK_BORDER)
    #Update List STAFF
    #Fill Truck
    sheetDB = wbSceduleDataBase.worksheets[1]
    sparks = SparksScheduleSearch()
    data = WeekScheduleExcelType()
    poolTable.Trucks = get_truck_distribution_data_base(filenameSceduleDataBase)
    data.Trucks = sparks.calcNewTrucks(poolTable)
    startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["column"]
    startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["row"]
    stepToTruck = 2
    j = 0
    while sheetDB.cell(row=startingPointRow + j, column=startingPointColumn).value != None:
        if data.Trucks[sheetDB.cell(row=startingPointRow + j, column=startingPointColumn).value] == None:
            sheetDB.cell(row=startingPointRow + j, column=startingPointColumn + stepToTruck).value = 0
        else:
            sheetDB.cell(row=startingPointRow + j, column=startingPointColumn + stepToTruck).value = data.Trucks[sheetDB.cell(row=startingPointRow + j, column=startingPointColumn).value]
        j += 1

    #Save DB's file
    wbSceduleDataBase.save(filenameSceduleDataBase)
    print(f"\nPool schedule number {numChoosingTimetable} was added to Schedule Data Base '{filenameSceduleDataBase}'")

def update_schedule_data_base_staff(filenameSceduleDataBase, poolOfNewStaff=None, numOfSelectedSchedule=-1):#TODO Потенциально не нужно
    #Check the 2nd parameter
    if poolOfNewStaff != None:
        print(f"{ERROR_STR_HEAD}! (update_schedule_data_base_staff)\n\t\tType of data wasn't selected")#Todo Add update from array of data
        return None
    #Init
    wbSceduleDataBase = openpyxl.load_workbook(filename=filenameSceduleDataBase)
    sheet = wbSceduleDataBase.worksheets[0]
    sheetStaff = wbSceduleDataBase.worksheets[1]
    lengthOfDataBase = sheet.cell(row=1, column=1).value
    # Check the 3rd parameter
    if numOfSelectedSchedule == -1:
        selectedNumber = lengthOfDataBase
    else:
        selectedNumber = numOfSelectedSchedule
    if ((selectedNumber < 1) or (selectedNumber > lengthOfDataBase)):
        if selectedNumber == 0:
            print(ERROR_STR_HEAD + "! (update_schedule_data_base_staff)\n\t\tEmpty data base!!!")  # it gives the last schedule in DB
        else:
            print(ERROR_STR_HEAD + "! (update_schedule_data_base_staff)\n\t\tOut of range!!!")  # it gives the last schedule in DB
        return None
    #Init table's coordinate
    tableWidth = 1 + 7 + SPACE_BETWEEN_TABLES  # +1 - begin from 1; 7 - week length; +1 - space between tables
    startingPointColumn = 1 + ((selectedNumber-1) * tableWidth)
    startingPointRow = 4
    staffPointColumn = 1
    staffPointRow = 2 #Because we already have HEAD
    i = 0
    statusChar = "-"
    #Table fill
    while sheet.cell(row=startingPointRow + i, column=startingPointColumn).value != None:
        formatting_cell(sheetStaff, staffPointRow + i, staffPointColumn, sheet.cell(row=startingPointRow+i, column=startingPointColumn).value,
                        FONT_SIZE, "Times New Roman", sheet.cell(row=startingPointRow + i, column=startingPointColumn).font.b, True, "right", "center")
        if sheet.cell(row=startingPointRow + i, column=startingPointColumn).font.b:
            statusChar = "+"
        else:
            statusChar = "-"
        formatting_cell(sheetStaff, staffPointRow + i, staffPointColumn + 1, statusChar,
                        FONT_SIZE, "Times New Roman", False, False, "center", "center")#IsElder
        formatting_cell(sheetStaff, staffPointRow + i, staffPointColumn + 2, 0,
                        FONT_SIZE, "Times New Roman", False, False, "center", "center")#NumOfTruck
        formatting_cell(sheetStaff, staffPointRow + i, staffPointColumn + 3, 3.0,
                        FONT_SIZE, "Times New Roman", False, False, "center", "center")  # NumOfHall
        sheetStaff.cell(row=staffPointRow + i, column=staffPointColumn).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)
        sheetStaff.cell(row=staffPointRow + i, column=staffPointColumn+1).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)
        sheetStaff.cell(row=staffPointRow + i, column=staffPointColumn+2).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)
        sheetStaff.cell(row=staffPointRow + i, column=staffPointColumn+3).border = openpyxl.styles.Border(right=THIN_BORDER, bottom=THIN_BORDER)
        i += 1
    #Save DB's file
    wbSceduleDataBase.save(filenameSceduleDataBase)
    print(f"\nStaff was updated on Schedule Data Base '{filenameSceduleDataBase}'")

#-GET_FUNCTIONS-
def get_schedule_list_data_base(filenameSceduleDataBase, numOfSelectedSchedule=-1):#TODO IsElder. How get status from table? Take it from Personal's sheet
    wbSceduleDataBase = openpyxl.load_workbook(filename=filenameSceduleDataBase)
    sheet = wbSceduleDataBase.worksheets[0]
    lengthOfDataBase = sheet.cell(row=1, column=1).value
    #Check parameters
    if numOfSelectedSchedule == -1:
        selectedNumber = lengthOfDataBase
    else:
        selectedNumber = numOfSelectedSchedule
    if ((selectedNumber < 1) or (selectedNumber > lengthOfDataBase)):
        if selectedNumber == 0:
            print(ERROR_STR_HEAD + "! (get_schedule_list_data_base)\n\t\tEmpty data base!!!")  # it gives the last schedule in DB
        else:
            print(ERROR_STR_HEAD + "! (get_schedule_list_data_base)\n\t\tOut of range!!!") #it gives the last schedule in DB
        outputEmptySchedule = WeekScheduleExcelType()
        return outputEmptySchedule
        # return None
    #Init tables parameters
    tableWidth = 1 + 7 + SPACE_BETWEEN_TABLES  # +1 - begin from 1; 7 - week length; +1 - space between tables
    startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["column"] + ((selectedNumber-1) * tableWidth)
    staffLength = 0
    r = 4
    while sheet.cell(row=r, column=startingPointColumn).value != None:
        staffLength += 1
        r += 1
    tableHeight = 1 + staffLength + SPACE_BETWEEN_TABLES
    startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List DB"]["row"]
    #Get selected schedule from data base
    try:
        outputSchedule = WeekScheduleExcelType()
        for i in range(1, tableHeight-SPACE_BETWEEN_TABLES):
            bufferName = sheet.cell(row=startingPointRow+i, column=startingPointColumn).value
            bufferIsElder = sheet.cell(row=startingPointRow+i, column=startingPointColumn).font.b
            bufferShifts = list[ShiftType]()
            for j in range(1, tableWidth-SPACE_BETWEEN_TABLES):
                if sheet.cell(row=startingPointRow+i, column=startingPointColumn+j).value != CHAR_CROSS:
                    if sheet.cell(row=startingPointRow+i, column=startingPointColumn+j).value.upper() == CHAR_TRUCK:
                        bufferShifts.append((j, 1.0, "Truck"))
                    elif sheet.cell(row=startingPointRow+i, column=startingPointColumn+j).value.upper() == CHAR_HALL:
                        bufferShifts.append((j, 1.0, "Hall"))
                    elif sheet.cell(row=startingPointRow+i, column=startingPointColumn+j).value.upper() == CHAR_HALF_HALL:
                        bufferShifts.append((j, 0.5, "Hall"))
            outputSchedule.EmployeeCards.append(EmployeeCard(name=bufferName, isElder=bufferIsElder, shifts=bufferShifts))
    except:
        print(f"{ERROR_STR_HEAD} (get_schedule_list_data_base)! Empty cell. Upper's error!")
        return 0
    else:
        return outputSchedule

def get_schedule_list_staff(filenameSceduleDataBase, numOfColumn = 1): #1 - NAME/IS_ELDER; 2 - NAME/NUM_OF_TRUCK; 3 - NAME/NUM_OF_HALL; 4 - NAME/PREFER_NUM_OF_HALL; 5 - NAME/UNDESIRABLE_DAYS.
    wbSceduleDataBase = openpyxl.load_workbook(filename=filenameSceduleDataBase)
    sheet = wbSceduleDataBase.worksheets[1]
    startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["column"]
    startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["STAFF_DATA"]["row"]
    stepToSelectColumn = numOfColumn

    if numOfColumn == 1:#1 - NAME/IS_ELDER;
        data = dict[str, bool]()
        j = 0
        while sheet.cell(row=startingPointRow + j, column=startingPointColumn).value != None:
            if sheet.cell(row=startingPointRow + j, column=startingPointColumn + stepToSelectColumn).value == "+":
                data[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = True
            else:
                data[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = False
            j += 1
    elif numOfColumn == 2:#2 - NAME/NUM_OF_TRUCK;
        data = dict[str, int]()
        j = 0
        while sheet.cell(row=startingPointRow + j, column=startingPointColumn).value != None:
            if sheet.cell(row=startingPointRow + j, column=startingPointColumn + stepToSelectColumn).value == None:
                data.Trucks[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = 0
            else:
                data.Trucks[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = sheet.cell(row=startingPointRow + j, column=startingPointColumn + stepToSelectColumn).value
            j += 1
        if j == 0:
            print(ERROR_STR_HEAD + "! (get_schedule_list_staff(numOfColumn = 2) Empty data base of trucks!")
            return None
    elif numOfColumn == 3 or numOfColumn == 4:#3 - NAME/NUM_OF_HALL; 4 - NAME/PREFER_NUM_OF_HALL;
        data = dict()#dict[str, int] for 3; dict[str, float] for 4
        j = 0
        while sheet.cell(row=startingPointRow + j, column=startingPointColumn).value != None:
            if sheet.cell(row=startingPointRow + j, column=startingPointColumn + stepToSelectColumn).value == None:
                data[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = 0
            else:
                data[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = sheet.cell(row=startingPointRow + j, column=startingPointColumn + stepToSelectColumn).value
            j += 1
    elif numOfColumn == 5:#5 - NAME/UNDESIRABLE_DAYS.
        data = dict()
        j = 0
        while sheet.cell(row=startingPointRow + j, column=startingPointColumn).value != None:
            undesirableDays = list()
            for i in range(WEEK_LENGTH):
                if sheet.cell(row=STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["row"] + j, column=STARTING_POINT_SCHEDULE_DATA_BASE["List STAFF"]["PREFER_OF_WEEK"]["column"] + i).value == CHAR_CROSS:
                    undesirableDays.append(i + 1)
            data[str(sheet.cell(row=startingPointRow + j, column=startingPointColumn).value)] = undesirableDays
            j += 1
    else:
        print(f"{ERROR_STR_HEAD} (get_schedule_list_staff)! Number of column is WRONG!")
        return None
    return data

def get_truck_distribution_data_base(filenameScheduleDataBase: str) -> TruckDistributionType:
    truck = get_schedule_list_staff(filenameScheduleDataBase, 2)
    shift = get_schedule_list_staff(filenameScheduleDataBase, 3)
    print(truck)
    print(shift)
    return {name: TruckElem(truck[name], shift[name]) for name in truck.keys()}

def get_schedule_list_coefficients(filenameSceduleDataBase) -> dict[str, float]:
    wbSceduleDataBase = openpyxl.load_workbook(filename=filenameSceduleDataBase)
    sheet = wbSceduleDataBase.worksheets[2]
    # """ Коэффициент дебатов для непрерывный череды смен""" self.shiftRepeatCoef = 10
    # """ Коэффициент дебатов для разницы между реальным количеством смен и желаемым для сотрудника""" self.differInShiftsCoef = 2.1
    # """ Коэффициент дебатов, когда сотрудник выходит на смену в не желаемый день""" self.undesirableDayCoef = 4
    listCoef = ["shiftRepeatCoef", "differInShiftsCoef", "undesirableDayCoef"]
    data = dict[str, float]()
    startingPointColumn = STARTING_POINT_SCHEDULE_DATA_BASE["List COEFFICIENTS"]["column"]
    startingPointRow = STARTING_POINT_SCHEDULE_DATA_BASE["List COEFFICIENTS"]["row"]
    i = 0
    for nameCoef in listCoef:
        data[nameCoef] = sheet.cell(row=startingPointRow+1, column=startingPointColumn+i).value
        i += 1
    # while sheet.cell(row=startingPointRow, column=startingPointColumn+i).value != None:
    #     data[sheet.cell(row=startingPointRow, column=startingPointColumn+i).value] = sheet.cell(row=startingPointRow+1, column=startingPointColumn+i).value
    #     i += 1
    return data
